"""
Caricamento del listone quotazioni.

Due formati supportati:
  - `load_quotazioni`: formato FantaMaster originale (Nome, Squadra, Ruolo,
    Quotazione). Mantenuto per compatibilità/debug.
  - `load_quotazioni_fvm`: formato Fantacalcio.it ufficiale con FVM (Fanta
    Valore di Mercato) — fonte primaria da questa versione in poi. Il file
    è tarato su un budget da 1000 crediti: FVM va dimezzato per la nostra
    lega da 500 (vedi build.py).
"""
from dataclasses import dataclass

import openpyxl


@dataclass
class QuotazioneRow:
    nome: str
    squadra: str
    ruolo: str
    quotazione: int


@dataclass
class QuotazioneFvmRow:
    nome: str
    squadra: str
    ruolo: str
    sub_ruolo: str | None
    quotazione: int  # Qt.A, quotazione attuale Classic
    quotazione_iniziale: int  # Qt.I
    fvm: int  # Fanta Valore di Mercato, base 1000 crediti


def load_quotazioni(path: str) -> list[QuotazioneRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Tutti"]

    rows: list[QuotazioneRow] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        nome, squadra, ruolo, quotazione = row[0], row[1], row[2], row[3]
        if not nome or not ruolo:
            continue  # righe di footer ("Ultimo aggiornamento...", ecc.)
        try:
            quotazione_int = int(quotazione)
        except (TypeError, ValueError):
            continue
        rows.append(
            QuotazioneRow(
                nome=str(nome).strip(),
                squadra=str(squadra).strip(),
                ruolo=str(ruolo).strip().upper(),
                quotazione=quotazione_int,
            )
        )
    return rows


def load_quotazioni_fvm(path: str) -> list[QuotazioneFvmRow]:
    """
    Legge il foglio "Tutti" del listone ufficiale Fantacalcio.it.
    Colonne: Id, R, RM, Nome, Squadra, Qt.A, Qt.I, Diff., Qt.A M, Qt.I M,
    Diff.M, FVM, FVM M.

    Il foglio "Ceduti" (giocatori non più in rosa alla squadra indicata)
    non viene incluso: non fanno parte del mercato d'asta.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Tutti"]

    rows: list[QuotazioneFvmRow] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        _id, ruolo, sub_ruolo, nome, squadra = row[0], row[1], row[2], row[3], row[4]
        qt_a, qt_i = row[5], row[6]
        fvm = row[11]
        if not nome or not ruolo:
            continue
        try:
            rows.append(
                QuotazioneFvmRow(
                    nome=str(nome).strip(),
                    squadra=str(squadra).strip(),
                    ruolo=str(ruolo).strip().upper(),
                    sub_ruolo=str(sub_ruolo).strip() if sub_ruolo else None,
                    quotazione=int(qt_a or 1),
                    quotazione_iniziale=int(qt_i or qt_a or 1),
                    fvm=int(fvm or 0),
                )
            )
        except (TypeError, ValueError):
            continue
    return rows
